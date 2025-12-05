#!/usr/bin/env python3
"""
Excel Column Selector
Reads an Excel file, allows interactive column selection, and saves a filtered version.
"""

import sys
import os
import pandas as pd
import inquirer
from openpyxl import load_workbook


def get_column_formats(filepath, column_names):
    """Extract number formats from the original Excel file for each column."""
    wb = load_workbook(filepath)
    ws = wb.active
    
    column_formats = {}
    for col_idx, col_name in enumerate(column_names, start=1):
        # Check format from first data row (row 2)
        cell = ws.cell(row=2, column=col_idx)
        if cell.number_format and cell.number_format != 'General':
            column_formats[col_name] = cell.number_format
    
    wb.close()
    return column_formats


def apply_column_formats(filepath, column_names, column_formats, row_count):
    """Apply number formats to the output Excel file."""
    wb = load_workbook(filepath)
    ws = wb.active
    
    for col_idx, col_name in enumerate(column_names, start=1):
        if col_name in column_formats:
            # Apply format to all data cells in this column (skip header)
            for row in range(2, row_count + 2):
                ws.cell(row=row, column=col_idx).number_format = column_formats[col_name]
    
    wb.save(filepath)
    wb.close()


def main():
    # Check command line arguments
    if len(sys.argv) != 2:
        print("Error: No Excel file specified!\n")
        print("Usage: python column_selector.py <excel_file_path>\n")
        print("Examples:")
        print("  python column_selector.py data.xlsx")
        print("  python column_selector.py /path/to/your/file.xlsx")
        print("  python column_selector.py ~/Documents/spreadsheet.xls\n")
        print("This tool will:")
        print("  1. Read your Excel file")
        print("  2. Let you select which columns to keep")
        print("  3. Save a filtered version with '_filtered' appended to the name")
        sys.exit(1)
    
    input_file = sys.argv[1]
    
    # Validate file exists
    if not os.path.exists(input_file):
        print(f"Error: File '{input_file}' not found.")
        sys.exit(1)
    
    # Validate file extension
    if not input_file.lower().endswith(('.xlsx', '.xls')):
        print("Error: File must be an Excel file (.xlsx or .xls)")
        sys.exit(1)
    
    print(f"Reading Excel file: {input_file}")
    
    try:
        # Read the Excel file
        df = pd.read_excel(input_file)
        
        print(f"\nFile loaded successfully! Found {len(df)} rows and {len(df.columns)} columns.\n")
        
        # Get available columns
        available_columns = list(df.columns)
        
        # Capture number formats from original file (for percentages, currency, etc.)
        column_formats = get_column_formats(input_file, available_columns)
        
        # Interactive column selection using checkboxes
        questions = [
            inquirer.Checkbox(
                'columns',
                message="Select columns to include (use spacebar to select, enter to confirm)",
                choices=available_columns,
                default=[]  # None selected by default
            )
        ]
        
        answers = inquirer.prompt(questions)
        
        # Check if user cancelled
        if answers is None or not answers['columns']:
            print("\nNo columns selected. Exiting.")
            sys.exit(0)
        
        selected_columns = answers['columns']
        
        print(f"\nSelected {len(selected_columns)} column(s): {', '.join(selected_columns)}")
        
        # Filter the dataframe
        filtered_df = df[selected_columns]
        
        # Generate output filename
        base_path = os.path.dirname(input_file)
        base_name = os.path.basename(input_file)
        name_without_ext, ext = os.path.splitext(base_name)
        output_file = os.path.join(base_path, f"{name_without_ext}_filtered{ext}")
        
        # Save the filtered Excel file
        print(f"\nSaving filtered file to: {output_file}")
        filtered_df.to_excel(output_file, index=False)
        
        # Apply original number formats (percentages, currency, etc.) to the output file
        apply_column_formats(output_file, selected_columns, column_formats, len(filtered_df))
        
        print(f"✓ Success! Filtered file saved with {len(filtered_df)} rows and {len(selected_columns)} columns.")
        
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()

